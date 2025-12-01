# yh_rag_cloud_api/parsers/budget_parser_excel_direct.py

import re
import logging
import pandas as pd
from typing import Dict, List, Any, Optional
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def parse_budget_excel_direct(file_path: str, sheet_name: str = "0") -> Dict[str, Any]:
    """
    Parse budget by reading Excel file directly with OpenPyXL.
    """
    try:
        # Load workbook
        wb = load_workbook(file_path, data_only=True)  # data_only=True to get calculated values

        # Get sheet
        if sheet_name.isdigit():
            sheet = wb.worksheets[int(sheet_name)]
        else:
            sheet = wb[sheet_name]

        metadata = {}
        budget_items = []

        # Extract metadata by scanning cells
        for row in sheet.iter_rows(values_only=True):
            for i, cell in enumerate(row):
                if cell and isinstance(cell, str):
                    cell_lower = cell.lower()

                    # Project Title
                    if "project title" in cell_lower:
                        if i + 1 < len(row) and row[i + 1]:
                            metadata["project_title"] = str(row[i + 1])

                    # Organization Name
                    elif "organisation name" in cell_lower or "organization name" in cell_lower:
                        if i + 1 < len(row) and row[i + 1]:
                            metadata["organization_name"] = str(row[i + 1])

                    # Funding Period
                    elif "funding period" in cell_lower:
                        if i + 1 < len(row) and row[i + 1]:
                            metadata["funding_period"] = str(row[i + 1])

                    # Total Amount Requested
                    elif "total amount requested" in cell_lower:
                        if i + 1 < len(row) and row[i + 1]:
                            amount = _extract_amount_direct(row[i + 1])
                            if amount:
                                metadata["total_amount_requested"] = amount

                    # Company ID
                    elif "company id" in cell_lower:
                        if i + 1 < len(row) and row[i + 1]:
                            metadata["organization_id"] = str(row[i + 1])

                    # Project ID
                    elif "project id" in cell_lower:
                        if i + 1 < len(row) and row[i + 1]:
                            metadata["project_id"] = str(row[i + 1])

            # Stop when we hit budget table headers
            if any(cell and "proposed budget" in str(cell).lower() for cell in row):
                break

        # Parse funding period dates
        if "funding_period" in metadata:
            from .budget_parser_cloud_2 import BudgetParserCloud2
            parser = BudgetParserCloud2()
            start_date, end_date = parser._parse_funding_period(metadata["funding_period"])
            metadata["funding_period_start"] = start_date
            metadata["funding_period_end"] = end_date

        # Extract budget items (similar to CSV method but with direct cell access)
        # ... implementation similar to CSV version but using sheet.cell() directly

        return {
            "metadata": metadata,
            "budget_items": budget_items,
            "parser_version": "excel_direct",
            "status": "success"
        }

    except Exception as e:
        logger.error(f"Error in direct Excel parser: {e}")
        return {
            "metadata": {},
            "budget_items": [],
            "parser_version": "excel_direct",
            "status": "error",
            "error": str(e)
        }


def _extract_amount_direct(value) -> Optional[float]:
    """Extract amount from direct Excel cell value."""
    if value is None:
        return None

    try:
        if isinstance(value, (int, float)):
            return float(value)
        elif isinstance(value, str):
            # Remove currency symbols and commas
            cleaned = re.sub(r'[RM\$,]', '', value.strip())
            return float(cleaned)
    except (ValueError, TypeError):
        return None

    return None